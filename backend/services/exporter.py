"""Excel exporter — export all DB tables to a single .xlsx workbook."""

import os
import logging
from datetime import datetime, timezone
from openpyxl import Workbook
from database import get_connection

logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")


def export_to_excel() -> str:
    """Export all relevant tables to an Excel workbook.

    Returns the file path of the generated .xlsx file.
    """
    conn = get_connection()
    wb = Workbook()

    # --- Jobs sheet ---
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"

    job_rows = conn.execute(
        """SELECT id, company, title, location, source, source_url, score,
                  status, date_found, date_applied, email_used, failure_step, notes
           FROM jobs ORDER BY date_found DESC"""
    ).fetchall()

    headers = ["ID", "Company", "Title", "Location", "Source", "URL", "Score",
               "Status", "Date Found", "Date Applied", "Email Used", "Failure Step", "Notes"]
    ws_jobs.append(headers)
    for row in job_rows:
        ws_jobs.append(list(row))

    # Auto-width columns
    for col in ws_jobs.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_jobs.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # --- Profile sheet ---
    ws_profile = wb.create_sheet("Profile")
    profile = conn.execute("SELECT * FROM profile WHERE id = 1").fetchone()
    if profile:
        for key in profile.keys():
            ws_profile.append([key, profile[key]])

    # --- Answer Bank sheet ---
    ws_answers = wb.create_sheet("Answer Bank")
    answer_rows = conn.execute("SELECT id, question_pattern, answer, category FROM answer_bank").fetchall()
    ws_answers.append(["ID", "Question Pattern", "Answer", "Category"])
    for row in answer_rows:
        ws_answers.append(list(row))

    # --- Assessments sheet ---
    ws_assess = wb.create_sheet("Assessments")
    assess_rows = conn.execute(
        """SELECT a.id, j.company, j.title, a.platform, a.oa_link,
                  a.deadline, a.status, a.received_date, a.notes
           FROM assessments a
           LEFT JOIN jobs j ON a.job_id = j.id
           ORDER BY a.deadline ASC"""
    ).fetchall()
    ws_assess.append(["ID", "Company", "Title", "Platform", "OA Link",
                       "Deadline", "Status", "Received", "Notes"])
    for row in assess_rows:
        ws_assess.append(list(row))

    # --- Emails sheet ---
    ws_emails = wb.create_sheet("Emails")
    email_rows = conn.execute(
        """SELECT id, job_id, from_address, subject, email_type,
                  received_date, action_needed
           FROM emails ORDER BY received_date DESC"""
    ).fetchall()
    ws_emails.append(["ID", "Job ID", "From", "Subject", "Type", "Received", "Action"])
    for row in email_rows:
        ws_emails.append(list(row))

    # --- Daily Stats sheet ---
    ws_stats = wb.create_sheet("Daily Stats")
    stat_rows = conn.execute(
        "SELECT * FROM daily_stats ORDER BY date DESC"
    ).fetchall()
    ws_stats.append(["Date", "Found", "Auto Applied", "Review", "Skipped",
                      "Duplicates", "Failed", "Responses"])
    for row in stat_rows:
        ws_stats.append(list(row))

    conn.close()

    # Save with timestamp
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"job_agent_export_{ts}.xlsx"
    filepath = os.path.join(DATA_DIR, filename)
    wb.save(filepath)

    logger.info(f"Excel export saved: {filepath}")
    return filepath


def aggregate_daily_stats():
    """Compute today's stats from jobs table and upsert into daily_stats."""
    conn = get_connection()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    stats = conn.execute(
        """SELECT
             COUNT(*) FILTER (WHERE date_found LIKE ?) as found,
             COUNT(*) FILTER (WHERE status = 'SUBMITTED' AND date_applied LIKE ?) as applied,
             COUNT(*) FILTER (WHERE status = 'REVIEW_NEEDED' AND date_found LIKE ?) as review,
             COUNT(*) FILTER (WHERE status = 'SKIPPED' AND date_found LIKE ?) as skipped,
             COUNT(*) FILTER (WHERE status = 'APPLY_FAILED' AND date_found LIKE ?) as failed
           FROM jobs""",
        (f"{today}%", f"{today}%", f"{today}%", f"{today}%", f"{today}%"),
    ).fetchone()

    if stats:
        conn.execute(
            """INSERT INTO daily_stats (date, jobs_found, auto_applied, review_queued, skipped, failed)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(date) DO UPDATE SET
                 jobs_found = excluded.jobs_found,
                 auto_applied = excluded.auto_applied,
                 review_queued = excluded.review_queued,
                 skipped = excluded.skipped,
                 failed = excluded.failed""",
            (today, stats[0], stats[1], stats[2], stats[3], stats[4]),
        )
        conn.commit()

    conn.close()
